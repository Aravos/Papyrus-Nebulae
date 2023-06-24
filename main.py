import logging
import os.path
import json
import csv
import re
from zipfile import ZipFile
import openpyxl

from adobe.pdfservices.operation.auth.credentials import Credentials
from adobe.pdfservices.operation.exception.exceptions import ServiceApiException, ServiceUsageException, SdkException
from adobe.pdfservices.operation.pdfops.options.extractpdf.extract_pdf_options import ExtractPDFOptions
from adobe.pdfservices.operation.pdfops.options.extractpdf.extract_element_type import ExtractElementType
from adobe.pdfservices.operation.execution_context import ExecutionContext
from adobe.pdfservices.operation.io.file_ref import FileRef
from adobe.pdfservices.operation.pdfops.extract_pdf_operation import ExtractPDFOperation

logging.basicConfig(level=os.environ.get("LOGLEVEL", "INFO"))
path = os.getcwd()

all_add = []
all_add2 = []
all_emails = []

def unzip(targetpath,extractpath):
    with ZipFile(targetpath, 'r') as zObject:
        zObject.extractall(extractpath)
        print('done')

def AdobeAPI(filename):
    file_exists = os.path.isfile(path + f"/Output/{filename}/structuredData.json")
    if not file_exists:
        print('NOT')
        # try:
        #     # get base path.
        #     base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

        #     # Initial setup, create credentials instance.
        #     credentials = Credentials.service_account_credentials_builder().from_file(path + "/pdfservices-api-credentials.json").build()

        #     # Create an ExecutionContext using credentials and create a new operation instance.
        #     execution_context = ExecutionContext.create(credentials)
        #     extract_pdf_operation = ExtractPDFOperation.create_new()

        #     # Set operation input from a source file.
        #     source = FileRef.create_from_local_file(path + "/InvoicesData/TestDataSet/"+filename+'.pdf')
        #     extract_pdf_operation.set_input(source)

        #     # Build ExtractPDF options and set them into the operation
        #     extract_pdf_options: ExtractPDFOptions = ExtractPDFOptions.builder() \
        #         .with_element_to_extract(ExtractElementType.TEXT) \
        #         .with_element_to_extract(ExtractElementType.TABLES) \
        #         .build()
        #     extract_pdf_operation.set_options(extract_pdf_options)

        #     result = extract_pdf_operation.execute(execution_context)
        #     result.save_as(path + f"/temp/{filename}.zip")
        #     unzip(path + "\\temp\\"+filename+".zip",path+"\\output\\"+filename)
        # except (ServiceApiException, ServiceUsageException, SdkException):
        #     #logging.exception("Exception encountered while executing operation")
        #     print("Exception encountered while executing operation")

def SetupCSV(csv_file_path):
    with open(csv_file_path, 'w', newline='') as file:
        writer = csv.writer(file)
        
        # Adding headers to CSV file
        headers = ['Bussiness__City',
                    'Bussiness__Country',
                    'Bussiness__Description',
                    'Bussiness__Name',
                    'Bussiness__StreetAddress',
                    'Bussiness__Zipcode',
                    'Customer__Address__line1',
                    'Customer__Address__line2',
                    'Customer__Email',
                    'Customer__Name',
                    'Customer__PhoneNumber',
                    'Invoice__BillDetails__Name',
                    'Invoice__BillDetails__Quantity',
                    'Invoice__BillDetails__Rate',
                    'Invoice__Description',
                    'Invoice__DueDate',
                    'Invoice__IssueDate',
                    'Invoice__Number',
                    'Invoice__Tax']
        writer.writerow(headers)


def toCSV(result):
    for i in result:
        print(i)
    
def Rows(filename):
    Bussiness__City = ""
    Bussiness__Country = ""
    Bussiness__Description = ""
    Bussiness__Name = ""
    Bussiness__StreetAddress = ""
    Bussiness__Zipcode = ""
    Customer__Address__line1 = ""
    Customer__Address__line2 = ""
    Customer__Email = ""
    Customer__Name = ""
    Customer__PhoneNumber = ""
    Invoice__Description = ""
    Invoice__DueDate = ""
    Invoice__IssueDate = ""
    Invoice__Number = ""
    Invoice__Tax = ""

    with open(path + f"/Output/{filename}/structuredData.json", 'r') as f:
        data = json.load(f)
        data = data.get('elements')
        removeLater = ""
        c = 0
        flg1 = 0
        flg2 = 0
        flg3 = 0
        Invoice = ''
        email = ''
        skip = 0
        add1 = 0
        add2 = 0
        flg100 = 0
        mm = 0
        duedate = 0
        duedate1 = 0
        for i in range(len(data)):

            if skip == 1:
                skip = 0
                continue
            if data[i].get('TextSize')==10.080001831054688:
                mm+=1
                #Get Customer Phone number using Wildcards

                if re.match('...-...-....',data[i].get('Text').strip()):
                    a = data[i].get('Text').strip().split(' ')
                    add1 = 1
                    c=0
                    while a[c] == '':
                        c+=1
                    Customer__PhoneNumber = a[c] 
                    continue

                if add1 == 1 and Customer__Address__line1 == '':
                    Customer__Address__line1 = data[i].get('Text')
                    for j in all_add:
                        if j[0] == Customer__Name:
                            Customer__Address__line1 = j[1]
                            break

                    if (Customer__Name,Customer__Address__line1) not in all_add:
                        all_add.append((Customer__Name,Customer__Address__line1))
                    add1 = 0
                    add2 = 1
                    continue
                
                if add2 == 1:
                    Customer__Address__line2 = data[i].get('Text')
                    for j in all_add2:
                        if j[0] == Customer__Name:
                            Customer__Address__line2 = j[1]
                            break

                    if (Customer__Name,Customer__Address__line2) not in all_add:
                        all_add2.append((Customer__Name,Customer__Address__line2))
                    add2 = 0
                    break

                #Getting Buissnes Names
                if c==0:
                    Bussiness__Name = data[i].get('Text')
                    c+=1
                elif c==1:
                    if flg1 == 0:
                        temp = data[i].get('Text')
                        temp = temp.split(',')
                        Bussiness__StreetAddress = temp[0]
                        Bussiness__City = temp[1]
                        if len(temp) == 4:
                            Bussiness__Country = temp[2] + ',' + temp[3]
                            c+=1
                        else:
                            flg1 = 1
                            continue
                    else:
                        Bussiness__Country = data[i].get('Text')
                        c+=1
        
                elif c==2:
                    Bussiness__Zipcode = data[i].get('Text')
                    c+=1


                elif c==3:
                    Invoice += data[i].get('Text')
                    if flg2 == 0:
                        # Invoice = data[i].get('Text')
                        temp = data[i].get('Text')
                        if 'Invoice# ' in temp:
                            temp = temp.replace('Invoice# ','*&*!@')
                        if 'Issue date ' in temp:
                            temp = temp.replace('Issue date ','*&*!@')
                        tempList = temp.split('*&*!@')
                        if len(tempList) == 2 and tempList[1]!='':
                            Invoice__Number = tempList[1]
                            flg3 = 1
                        if len(tempList) == 3 and tempList[2]!='':
                            Invoice__Number = tempList[1]
                            Invoice__IssueDate = tempList[2]
                            c+=1
                        else:
                            flg2 = 1
                            continue
                    else:
                        if flg3 == 0:
                            # Invoice += data[i].get('Text')
                            temp = Invoice
                            if 'Invoice# ' in temp:
                                temp = temp.replace('Invoice# ','*&*!@')
                            if 'Issue date ' in temp:
                                temp = temp.replace('Issue date ','*&*!@')
                            tempList = temp.split('*&*!@')
                            if len(tempList) >=2 and tempList[1]!='':
                                Invoice__Number = tempList[1]
                                if tempList[2] == '':
                                    flg3 = 1
                                else:
                                    c+=1
                        else:
                            temp =  data[i].get('Text').strip()
                            if re.match('..-..-....',temp):
                                Invoice__IssueDate = data[i].get('Text').strip()
                                c+=1

                elif c==4:
                    Bussiness__Description = data[i].get('Text')
                    c+=1
                elif c==5:
                    if data[i].get('Text').lower().strip() not in  ['bill to','details','payment']:
                        if '@' not in data[i].get('Text'):
                            Customer__Name = data[i].get('Text')
                        else:
                            tempList = data[i].get('Text').split(' ')
                            Customer__Name = tempList[0] + ' ' + tempList[1]
                            if len(tempList)>=4:
                                if tempList[3] != '':
                                    Customer__Email = tempList[2]+tempList[3]
                                    all_emails.append(Customer__Email)
                                else:
                                    Customer__Email+=tempList[2]
                                    for i in all_emails:
                                        if Customer__Email in i:
                                            Customer__Email = i
                                            break
                                    if Customer__Email==tempList[2]:
                                        counter = 1
                                        while data[i+counter].get('Text') == None:
                                            counter+=1
                                        Customer__Email+=data[i+counter].get('Text')
                            if len(tempList)>=5:
                                Customer__PhoneNumber = tempList[4]

                            if len(tempList)>=7:
                                for i in tempList[5:len(tempList)-3]:
                                    Customer__Address__line1 += i + ' '
                                Customer__Address__line1 = Customer__Address__line1.strip()
                                for i in tempList[len(tempList)-3:]:
                                    Customer__Address__line2 += i + ' '
                                Customer__Address__line2 =  Customer__Address__line2.strip()
                        c+=1
                elif c==6:
                    if Customer__Email == '':
                        if '@' in data[i].get('Text'):
                            if '.com' not in data[i].get('Text'):
                                if data[i+1].get('Text') != None:
                                    Customer__Email = data[i].get('Text').strip()+data[i+1].get('Text').strip()
                                    skip = 1
                                    c+=1
                                    continue
                                else:
                                    Customer__Email = data[i].get('Text').strip()
                                    for j in all_emails:
                                        if Customer__Email in j:
                                            Customer__Email = j
                                            break
                                    if Customer__Email == data[i].get('Text').strip():
                                        counter = 1
                                        while data[i+counter].get('Text') == None:
                                            counter+=1
                                        Customer__Email+=data[i+counter].get('Text')
                                    c+=1
                            else:
                                Customer__Email = data[i].get('Text').strip()
                                c+=1
                        continue
                    
                    else:
                        c+=1
                

        
        Bussiness__StreetAddress = Bussiness__StreetAddress.strip()
        Bussiness__City = Bussiness__City.strip()
        Bussiness__Country = Bussiness__Country.strip()
        if Invoice__IssueDate!='':pass
        else:
            Invoice = Invoice.strip().split(' ')
            if re.match('..-..-....',Invoice[-1]):
                Invoice__IssueDate = Invoice[-1]
        
        for i in data:
            if i.get('Text'):
                if "Due date".lower() in i.get('Text').strip().lower():
                    Invoice__DueDate = i.get('Text').strip().split(' ')[2]
                    break
        mm = -10
        ff = 0
        while 1:
            if data[mm].get('Text')!=None:
                if 'Tax %' in data[mm].get('Text').strip():
                    if len(data[mm].get('Text').strip().split(" "))>2:
                        ff = 1
                        break 
                if data[mm].get('Text').strip().isnumeric() and len(data[mm].get('Text').strip()) <= 2:
                    break
            if mm == -1:
                break
            mm+=1
        if ff == 1:
            Invoice__Tax = data[mm].get('Text').strip().split(' ')[2]
        else:
            Invoice__Tax = data[mm].get('Text').strip()
        if Invoice__Description == '':
            file_exists = os.path.isfile(path+f'/Output/{filename}/tables/fileoutpart0.xlsx')
            if file_exists:
                workbook = openpyxl.load_workbook(path+f'/Output/{filename}/tables/fileoutpart0.xlsx')
                worksheet = workbook['Sheet1']
                des = ''
                for row in worksheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell!=None:
                            des += cell.replace(' _x000D_','') + ' '
                        break
                des = des.replace('  ',' ').replace('DETAILS','')
                workbook.close()
                if 'QTY' in des or 'ITEM' in des or 'BILL TO' in des:
                    des = ''
            if des == '':
                for i in range(len(data)):
                    a = data[i].get('Text')
                    if a!=None and data[i].get('Page') == 0:
                        if "details" in a.lower():
                            i+=1
                            try:
                                while 1:
                                    b = str(data[i].get('Text'))
                                    a = data[i].get('Text')
                                    i+=1
                                    if b!=None:
                                        b = b.replace(' ','')
                                        if 'ITEM' in b:
                                            b = ''
                                            break
                                        if 'Duedate:' not in b and 'PAYMENT' not in b and Customer__Name.replace(' ','') not in b and Customer__Email not in b and Customer__PhoneNumber not in b and Customer__Address__line1.replace(' ','') not in b and Customer__Address__line2.replace(' ','') not in b and Customer__Email.replace(' ','') not in b and '@' not in b and '$' not in b and 'il.com' not in b and len(b.strip())>1 and 'None' not in b:
                                            break
                                Invoice__Description = b
                            except:
                                Invoice__Description = ''
                            break
                    
        if Customer__Address__line1 == '':
            for i in range(len(data)):
                a = data[i].get('Text')
                if a != None:
                    if Customer__PhoneNumber in a:
                        a = data[i+1].get('Text')
                        if a!=None:
                            a = a.strip().split(' ')
                            Customer__Address__line2 = a[-1]
                            for i in range(len(a)-1):
                                Customer__Address__line1 += a[i] + ' '
                            Customer__Address__line1 = Customer__Address__line1.strip()
                            break

        l =[Bussiness__City,
            Bussiness__Country,
            Bussiness__Description,
            Bussiness__Name,
            Bussiness__StreetAddress,
            Bussiness__Zipcode,
            Customer__Address__line1,
            Customer__Address__line2,
            Customer__Email,
            Customer__Name,
            Customer__PhoneNumber,
            Invoice__Description,
            Invoice__DueDate,
            Invoice__IssueDate,
            Invoice__Number,Invoice__Tax]
        
        i = 2
        while 1:
            try:
                workbook = openpyxl.load_workbook(path+f'/Output/{filename}/tables/fileoutpart{i}.xlsx')
                break
            except FileNotFoundError:
                i-=1
                pass
        alert = 0
        csv_file_path = path + '/MyExtractedData.csv'
        worksheet = workbook['Sheet1']
        with open(csv_file_path, 'a', newline='') as file:
            writer = csv.writer(file)
            for row in worksheet.iter_rows(values_only=True):
                if alert == 0:
                    j = [0,0,0]
                    c = 0
                    for cell in row:
                        j[c] = cell.replace(' _x000D_','')
                        c+=1
                        if c==3:
                            break
                        if 'Subtotal' in str(cell):
                            alert = 1
                            workbook.close()
                            break
                    if alert == 1:
                        continue
                    writer.writerow(l[:11]+j+l[11:])
        workbook.close()
        if alert == 1:
            i-=1
            workbook = openpyxl.load_workbook(path+f'/Output/{filename}/tables/fileoutpart{i}.xlsx')
            worksheet = workbook['Sheet1']
            with open(csv_file_path, 'a', newline='') as file:
                writer = csv.writer(file)
                for row in worksheet.iter_rows(values_only=True):
                    j = [0,0,0]
                    c = 0
                    for cell in row:
                        j[c] = cell.replace(' _x000D_','')
                        c+=1
                        if c==3:
                            break
                    writer.writerow(l[:11]+j+l[11:])
                    alert = 0
        workbook.close()


    

#   DRIVER CODE
csv_file_path = path + '/MyExtractedData.csv'
SetupCSV(csv_file_path)
for i in range(0,100):
    AdobeAPI('output'+str(i))
    Rows('output'+str(i))